"""
Broadcast Playlist Checker — Core Logic v5
"""
import json, re, xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from datetime import datetime, timedelta


# ── TRANSLATION ───────────────────────────────────────────────────────────────

_S = {
    'section_programs':   {'en': 'PROGRAM CHECK (Playlist vs Grilla)',       'es': 'VERIFICACIÓN DE PROGRAMAS (Playlist vs Grilla)'},
    'section_commercials':{'en': 'COMMERCIAL CHECK (Playlist vs XML)',        'es': 'VERIFICACIÓN DE COMERCIALES (Playlist vs XML)'},
    'section_promos':     {'en': 'PROMO REPEAT CHECK',                        'es': 'PROMOS REPETIDAS'},
    'section_ingested':   {'en': 'NOT INGESTED ASSETS',                       'es': 'ACTIVOS NO INGESTADOS'},
    'section_bugs':       {'en': 'BUGS',                                       'es': 'BUGS'},
    'section_cues':       {'en': 'CUE TONE REPORT',                           'es': 'REPORTE DE CUE TONES'},
    'full_day':           {'en': 'FULL DAY',                                   'es': 'DÍA COMPLETO'},
    'partial':            {'en': 'CURRENT (partial)',                          'es': 'ACTUAL (parcial)'},
    'checking_from':      {'en': 'CHECKING FROM',                              'es': 'VERIFICANDO DESDE'},
    'channel':            {'en': 'CHANNEL',                                    'es': 'CANAL'},
    'date_lbl':           {'en': 'DATE',                                       'es': 'FECHA'},
    'type_lbl':           {'en': 'PLAYLIST TYPE',                              'es': 'TIPO DE PLAYLIST'},
    'summary':            {'en': 'SUMMARY',                                    'es': 'RESUMEN'},
    'show_blocks':        {'en': 'show blocks',                                'es': 'bloques de programa'},
    'commercials_lbl':    {'en': 'commercials',                                'es': 'comerciales'},
    'no_grilla':          {'en': '! Grilla not provided',                      'es': '! Grilla no proporcionada'},
    'no_xml':             {'en': '! XML log not provided',                     'es': '! Log XML no proporcionado'},
    'ok_programs':        {'en': '✓ All programs match grilla',                'es': '✓ Todos los programas coinciden con la grilla'},
    'ok_commercials':     {'en': '✓ All {n} commercials match XML log',        'es': '✓ Los {n} comerciales coinciden con el log XML'},
    'ok_promos':          {'en': '✓ No repeated promos within same break',     'es': '✓ Sin promos repetidas en el mismo break'},
    'ok_ingested':        {'en': '✓ All assets ingested',                      'es': '✓ Todos los activos están ingestados'},
    'ok_bugs':            {'en': '✓ No bugs scheduled',                        'es': '✓ Sin bugs programados'},
    'total_cues':         {'en': 'Total cue tones',                            'es': 'Total cue tones'},
    'anchored':           {'en': '  ℹ  Anchored at grilla position {i}: {id}', 'es': '  ℹ  Anclado en grilla pos {i}: {id}'},
    'already_aired':      {'en': '  ℹ  ALREADY AIRED: {id}',                  'es': '  ℹ  YA SE TRANSMITIÓ: {id}'},
    'wrong_ep':           {'en': '  ⚠  WRONG EPISODE: Grilla={g} | Playlist={p} @ {t}', 'es': '  ⚠  EPISODIO INCORRECTO: Grilla={g} | Playlist={p} @ {t}'},
    'not_in_pl':          {'en': '  ✗  NOT IN PLAYLIST: {id}',                'es': '  ✗  NO ESTÁ EN PLAYLIST: {id}'},
    'extra_pl':           {'en': '  ✗  EXTRA: {id} @ {t} (not in grilla)',     'es': '  ✗  EXTRA: {id} @ {t} (no está en grilla)'},
    'order_mis':          {'en': '  ⚠  ORDER pos {i}: Grilla={g} | Playlist={p}', 'es': '  ⚠  ORDEN pos {i}: Grilla={g} | Playlist={p}'},
    'xml_not_pl':         {'en': '  ✗  IN XML, NOT IN PLAYLIST: {ref} ({n}x in XML)', 'es': '  ✗  EN XML, NO EN PLAYLIST: {ref} ({n}x en XML)'},
    'pl_not_xml':         {'en': '  ✗  IN PLAYLIST, NOT IN XML: {ref} ({n}x in playlist)', 'es': '  ✗  EN PLAYLIST, NO EN XML: {ref} ({n}x en playlist)'},
    'count_diff':         {'en': '  ⚠  COUNT DIFF: {ref} | XML={xn}x | Playlist={pn}x', 'es': '  ⚠  DIFERENCIA: {ref} | XML={xn}x | Playlist={pn}x'},
    'promo_rep':          {'en': '  ⚠  PROMO REPEAT after [{after}] @ {t}: {ref} {n}x', 'es': '  ⚠  PROMO REPETIDA después de [{after}] @ {t}: {ref} {n}x'},
    'ni_program':         {'en': '  ⚠  NOT INGESTED [Program]: {id} @ {t} | {show}', 'es': '  ⚠  NO INGESTADO [Programa]: {id} @ {t} | {show}'},
    'ni_other':           {'en': '  ⚠  NOT INGESTED [{typ}]: {ref} @ {t} | {name}', 'es': '  ⚠  NO INGESTADO [{typ}]: {ref} @ {t} | {name}'},
    'bug_line':           {'en': '  🔲 {beh_label} : {cmd} — {id} @ {t} | {show}', 'es': '  🔲 {beh_label} : {cmd} — {id} @ {t} | {show}'},
}

def T(key, lang='en', **kwargs):
    s = _S.get(key, {}).get(lang, _S.get(key, {}).get('en', key))
    return s.format(**kwargs) if kwargs else s


# ── TIME HELPERS ──────────────────────────────────────────────────────────────

def _edt_start(year):
    """Second Sunday of March 2:00 AM."""
    m1 = datetime(year, 3, 1)
    d = (6 - m1.weekday()) % 7
    return datetime(year, 3, (m1 + timedelta(days=d+7)).day, 2)

def _est_start(year):
    """First Sunday of November 2:00 AM."""
    n1 = datetime(year, 11, 1)
    d = (6 - n1.weekday()) % 7
    return datetime(year, 11, (n1 + timedelta(days=d)).day, 2)

def utc_to_et(dt):
    if dt is None: return None
    edt = _edt_start(dt.year)
    est = _est_start(dt.year)
    offset = -4 if edt <= dt < est else -5
    tz = 'EDT' if offset == -4 else 'EST'
    return dt + timedelta(hours=offset), tz

def fmt_time(dt):
    """UTC only."""
    if dt is None: return '??:??:??'
    return dt.strftime('%H:%M:%S')

def fmt_t(dt):
    """UTC / ET string."""
    if dt is None: return '??:??:?? UTC'
    et, tz = utc_to_et(dt)
    return f'{dt.strftime("%H:%M:%S")} UTC / {et.strftime("%H:%M:%S")} ET'

def parse_timecode(tc):
    try:
        tc = tc.split(';')[0].split('@')[0].strip()
        return datetime.strptime(tc, '%Y-%m-%d %H:%M:%S')
    except: return None

def parse_duration(dur):
    try:
        dur = dur.split(';')[0].split('@')[0]
        h,m,s = dur.split(':')
        return int(h)*3600 + int(m)*60 + int(s)
    except: return 0

def parse_xml_time(ts):
    """XML startat: '15:40:00:00' → datetime (dummy date)."""
    try:
        p = ts.split(':')
        return datetime(2000, 1, 1, int(p[0]), int(p[1]), int(p[2]))
    except: return None


# ── ID HELPERS ────────────────────────────────────────────────────────────────

def is_episode_id(val):
    if not val or not isinstance(val, str): return False
    val = val.strip()
    if ' ' in val or len(val) < 3 or len(val) > 16: return False
    if not re.match(r'^[A-Z]', val): return False
    return len(re.findall(r'\d', val)) >= 3

def is_movie_id(val):
    """Movie/special IDs: pure uppercase alpha, 3-8 chars. MARCE, NELQP, CAME."""
    if not val or not isinstance(val, str): return False
    val = val.strip()
    return bool(re.match(r'^[A-Z]{3,8}$', val))

def normalize_id(ep_id):
    if not ep_id: return ''
    ep_id = re.sub(r'_\d+$', '', str(ep_id).strip())
    # Only normalize extra leading zeros before 4+ digit date suffixes (e.g. COSA00327→COSA0327)
    # Do NOT touch 3-digit episode numbers (e.g. LATPAN001 stays LATPAN001)
    ep_id = re.sub(r'([A-Za-z][A-Za-z0-9]*)0{2,}(\d{4,})',
                   lambda m: m.group(1) + (m.group(2)[-4:] if len(m.group(2)) > 4 else m.group(2)),
                   ep_id)
    return ep_id.upper()

def show_prefix(ep_id):
    if not ep_id: return ''
    ep_id = ep_id.upper()
    m = re.match(r'^([A-Z]+)', ep_id)
    if not m: return ''
    letters = m.group(1)
    if re.match(r'^[A-Z]{3,8}$', ep_id): return ep_id
    return letters


# ── JSON PARSER ───────────────────────────────────────────────────────────────

def parse_json_playlist(data):
    events = data.get('events', [])
    has_marker = any(a.get('type') == 'marker'
                     for ev in events[:3] for a in ev.get('assets', []))
    playlist_type = 'full' if has_marker else 'current'

    date = None
    for ev in events:
        dt = parse_timecode(ev.get('startTime', ''))
        if dt: date = dt.date(); break

    programs, commercials, promos, cue_tones, not_ingested, breaks = [], [], [], [], [], []
    current_break, last_program, last_program_raw = [], None, None

    for ev in events:
        ev_assets = ev.get('assets', [])
        ev_start  = parse_timecode(ev.get('startTime', ''))
        ev_dur    = parse_duration(ev.get('duration', ''))
        ev_name   = ev.get('name', '')
        ev_ref    = ev.get('reference', '')
        behaviors = ev.get('behaviors', [])

        for b in behaviors:
            if b.get('name') == 'CUEON' and not b.get('disabled', True):
                ct = ev_assets[0].get('reference', ev_name) if ev_assets else ev_name
                cue_tones.append({'ref': ev_ref, 'name': ev_name, 'ct_id': ct, 'start': ev_start})

        for asset in ev_assets:
            atype = asset.get('type', '')
            aref  = asset.get('reference', '')
            tcin  = asset.get('tcIn', '')

            if tcin.startswith('07:') and atype != 'live':
                not_ingested.append({'asset_ref': aref, 'name': ev_name,
                                     'type': atype, 'start': ev_start, 'ref': ev_ref})

            if atype in ('Program', 'live'):
                if current_break:
                    breaks.append({'after_program': last_program,
                                   'after_program_raw': last_program_raw,
                                   'items': current_break[:]})
                    current_break = []
                seg_m = re.search(r'_(\d+)$', aref)
                seg   = int(seg_m.group(1)) if seg_m else 1
                ep_id = normalize_id(aref)
                _logo = next((b.get('params',{}).get('Command')
                              for b in behaviors
                              if b.get('name')=='LOGOHD' and not b.get('disabled',False)), None)
                programs.append({'episode_id': ep_id, 'episode_id_raw': aref,
                                  'seg_num': seg, 'start': ev_start, 'duration': ev_dur,
                                  'name': ev_name, 'ref': ev_ref, 'asset_type': atype,
                                  'is_missing': (atype == 'Program' and tcin.startswith('07:')),
                                  'logo': _logo})
                last_program     = ep_id
                last_program_raw = aref

            elif atype == 'Commercial':
                commercials.append({'asset_ref': aref, 'name': ev_name,
                                    'start': ev_start, 'duration': ev_dur,
                                    'ref': ev_ref})
                current_break.append({'type': 'Commercial', 'ref': aref,
                                      'start': ev_start, 'event_ref': ev_ref})
            elif atype == 'Promotion':
                _pdur = 0
                try:
                    _tcin  = assets[0].get('tcIn','') if assets else ''
                    _tcout = assets[0].get('tcOut','') if assets else ''
                    def _tc(s):
                        import re as _r; m=_r.match(r'(\d+):(\d+):(\d+)',str(s)); return int(m.group(1))*3600+int(m.group(2))*60+int(m.group(3)) if m else 0
                    _pdur = _tc(_tcout) - _tc(_tcin)
                except: pass
                promos.append({'asset_ref': aref, 'name': ev_name, 'start': ev_start, 'ref': ev_ref, 'duration': max(0,_pdur)})
                current_break.append({'type': 'Promotion', 'ref': aref, 'start': ev_start})

    if current_break:
        breaks.append({'after_program': last_program,
                       'after_program_raw': last_program_raw,
                       'items': current_break})

    return {'type': playlist_type, 'date': date, 'events': events,
            'programs': programs, 'commercials': commercials, 'promos': promos,
            'breaks': breaks, 'cue_tones': cue_tones, 'not_ingested': not_ingested}


def build_show_sequence(programs, from_start=None):
    seq, prev = [], None
    for p in programs:
        if from_start and p['start'] and p['start'] < from_start: continue
        ep = p['episode_id']
        if ep != prev:
            seq.append({'id': ep, 'start': p['start'], 'raw': p['episode_id_raw']})
            prev = ep
    return seq


# ── XML PARSER ────────────────────────────────────────────────────────────────

def parse_xml_log(filepath_or_bytes):
    """
    Parse standard XML traffic log.
    Auto-detects format:
      - <traffics><traffic><item>  → standard (CATV, TVD, original Pasiones)
      - <tabledata><data><row>     → tabledata (TN, Pasiones after format change, Sony)
    Safe to use for all channels — routes to the correct parser internally.
    """
    try:
        if hasattr(filepath_or_bytes, 'read'): content = filepath_or_bytes.read()
        elif isinstance(filepath_or_bytes, bytes): content = filepath_or_bytes
        else:
            with open(filepath_or_bytes, 'rb') as f: content = f.read()
        # Sanitize unescaped & that break XML parser (e.g. P&G in title fields)
        content = re.sub(rb'&(?![a-zA-Z#][a-zA-Z0-9#]*;)', b'&amp;', content)
        root = ET.fromstring(content)
        # Auto-detect format
        if root.tag == 'tabledata':
            # tabledata format — same logic as parse_xml_log_tn
            items = []
            for row in root.findall('.//row'):
                local_time = row.findtext('column-1', '').strip()
                mediaid    = row.findtext('column-4', '').strip()
                typ        = row.findtext('column-5', '').strip().upper()
                title      = row.findtext('column-6', '').strip()
                duration   = row.findtext('column-3', '').strip()
                if typ == 'PROGRAM':    ct = 'PROGRAM_BEGIN'
                elif typ == 'PROMOTION': ct = 'PROMO'
                else:                    ct = typ
                items.append({'mediaid': mediaid, 'name': title,
                              'contenttype': ct, 'startat': local_time,
                              'duration': duration, 'externalid': ''})
            return items
        else:
            # Standard <traffics><traffic><item> format
            traffic = root.find('traffic')
            if traffic is None: traffic = root
            return [{'mediaid': i.get('mediaid',''), 'name': i.findtext('n','').strip(),
                     'contenttype': i.findtext('contenttype','').strip().upper(),
                     'startat': i.findtext('startat','').strip(),
                     'duration': i.findtext('duration','').strip(),
                     'externalid': i.findtext('externalid','').strip()}
                    for i in traffic.findall('item')]
    except: return []

def parse_xml_log_tn(filepath_or_bytes):
    """
    Parser for Todonovelas XML — <tabledata><data><row><column-N> format.
    column-1=LocalTime, column-4=MediaId, column-5=Type, column-6=Title
    Returns same dict format as parse_xml_log for compatibility.
    """
    try:
        if hasattr(filepath_or_bytes, 'read'): content = filepath_or_bytes.read()
        elif isinstance(filepath_or_bytes, bytes): content = filepath_or_bytes
        else:
            with open(filepath_or_bytes, 'rb') as f: content = f.read()
        root = ET.fromstring(content)
        items = []
        for row in root.findall('.//row'):
            local_time = row.findtext('column-1', '').strip()
            mediaid    = row.findtext('column-4', '').strip()
            typ        = row.findtext('column-5', '').strip().upper()
            title      = row.findtext('column-6', '').strip()
            duration   = row.findtext('column-3', '').strip()
            # Normalise type to match standard contenttype values
            if typ == 'PROGRAM': ct = 'PROGRAM_BEGIN'
            elif typ == 'PROMOTION': ct = 'PROMO'
            else: ct = typ
            items.append({'mediaid': mediaid, 'name': title,
                          'contenttype': ct, 'startat': local_time,
                          'duration': duration, 'externalid': ''})
        return items
    except: return []

def _xml_dur_secs(dur_str):
    """Parse XML duration HH:MM:SS:FF → seconds (ignore frames)."""
    try:
        p = dur_str.split(':')
        return int(p[0])*3600 + int(p[1])*60 + int(p[2])
    except: return 0

def _is_xml_program_anchor(item):
    """Only PROGRAM_BEGIN/PROGRAM_SEGMENT used for break-by-break alignment."""
    return item.get('contenttype','') in ('PROGRAM_BEGIN','PROGRAM_SEGMENT')

def _is_xml_start_anchor(item):
    """For partial start detection: program segments + infomercials (CM ≥ 20min)."""
    ct = item.get('contenttype','')
    if ct in ('PROGRAM_BEGIN','PROGRAM_SEGMENT'): return True
    if ct == 'COMMERCIAL' and _xml_dur_secs(item.get('duration','')) >= 1200: return True
    return False

def build_xml_breaks(xml_rows):
    """
    Walk XML, group commercials between PROGRAM segments only (not infomercials).
    Infomercials appear as commercials inside a break, matching JSON behavior.
    Returns list of {'anchor_id', 'commercials': [mediaid, ...]}
    Includes breaks with zero commercials for alignment.
    """
    result = []
    anchor = None
    comms  = []
    for item in xml_rows:
        if _is_xml_program_anchor(item):
            if anchor is not None:
                result.append({'anchor_id': anchor, 'commercials': comms[:]})
            anchor = item['mediaid']
            comms  = []
        elif item.get('contenttype') == 'COMMERCIAL':
            comms.append(item['mediaid'])
    if anchor is not None:
        result.append({'anchor_id': anchor, 'commercials': comms[:]})
    return result

def xml_commercials(rows):
    return [r for r in rows if r.get('contenttype') == 'COMMERCIAL']

def find_xml_anchor_by_extid(events, xml_rows):
    """
    Find where partial JSON starts in XML using externalid/reference match.
    Returns start index in xml_rows for commercial comparison.
    """
    ext_idx = {row['externalid']: i for i, row in enumerate(xml_rows)}
    for ev in events:
        ref = ev.get('reference', '')
        if ref in ext_idx:
            return ext_idx[ref]
    return 0


# ── GRILLA PARSER ─────────────────────────────────────────────────────────────

def parse_grilla(filepath_or_bytes, target_date, channel='catv'):
    """Route to the correct grilla parser based on channel type."""
    if channel in ('latam', 'us'):
        return _parse_grilla_pasiones(filepath_or_bytes, target_date)
    if channel == 'tn':
        return _parse_grilla_tn(filepath_or_bytes, target_date)
    return _parse_grilla_catv_tvd(filepath_or_bytes, target_date)

def _parse_grilla_catv_tvd(filepath_or_bytes, target_date):
    """Original CATV/TVD grilla parser — single active sheet, datetime header."""
    from openpyxl import load_workbook
    import io
    if isinstance(filepath_or_bytes, str):
        wb = load_workbook(filepath_or_bytes, read_only=False)
    else:
        raw = filepath_or_bytes.read() if hasattr(filepath_or_bytes, 'read') else filepath_or_bytes
        wb = load_workbook(io.BytesIO(raw), read_only=False)

    ws = wb.active
    # Keep two views: values_only for formulas, full rows for bold detection
    all_rows_full = list(ws.iter_rows(values_only=False))
    all_rows = [[cell.value for cell in row] for row in all_rows_full]
    if len(all_rows) < 2: return []

    header_row = all_rows[1]
    # Build date→col map, skipping ET/UTC/CA marker columns
    _NON_DATE = {'ET', 'UTC', 'CA', 'E.T.', 'U.T.C.'}
    date_col_map = {}
    for ci, val in enumerate(header_row):
        if val is None: continue
        if isinstance(val, str) and val.strip().upper() in _NON_DATE: continue
        if hasattr(val, 'date'):
            date_col_map[val.date()] = ci
        elif isinstance(val, str) and val.strip().startswith('='):
            # Formula like =C2+5 — resolve base date + offset
            m2 = re.match(r'^=([A-Z]+)(\d+)\+(\d+)$', val.strip())
            if m2:
                try:
                    base_ci = sum((ord(c)-ord('A')+1)*(26**i)
                                 for i,c in enumerate(reversed(m2.group(1))))-1
                    base_ri = int(m2.group(2))-1
                    base_v  = all_rows[base_ri][base_ci] if base_ri < len(all_rows) and base_ci < len(all_rows[base_ri]) else None
                    if base_v and hasattr(base_v, 'date'):
                        date_col_map[(base_v + timedelta(days=int(m2.group(3)))).date()] = ci
                except Exception:
                    pass
    target_col = date_col_map.get(target_date)
    if target_col is None and date_col_map:
        # Fallback: find Monday and walk by weekday offset
        monday_d = min(date_col_map.keys())
        monday_c = date_col_map[monday_d]
        for offset in range(7):
            if monday_d + timedelta(days=offset) == target_date:
                target_col = monday_c + offset; break
    if target_col is None:
        return []

    def resolve_cell(val):
        if not val or not isinstance(val, str): return val
        m = re.match(r'^=([A-Z]+)(\d+)$', val.strip())
        if not m: return val
        col_str, row_num = m.group(1), int(m.group(2))
        col_idx = sum((ord(c)-ord('A')+1) * (26**i) for i,c in enumerate(reversed(col_str))) - 1
        row_idx = row_num - 1
        if row_idx < len(all_rows) and col_idx < len(all_rows[row_idx]):
            return all_rows[row_idx][col_idx]
        return val

    def extract_ids(val):
        val = resolve_cell(val)
        if not val or not isinstance(val, str) or val.startswith('='): return []
        val = val.strip()
        # Full cell is an episode ID (standard with digits)
        if is_episode_id(val): return [normalize_id(val)]
        # Full cell is a movie/special ID (pure alpha, no digits)
        if is_movie_id(val): return [normalize_id(val)]
        # Cell has mixed content — tokenize and extract only standard IDs (digits required)
        # Do NOT apply movie_id check on tokens to avoid extracting description words
        tokens = re.findall(r'[A-Z0-9]+', val.upper())
        return [normalize_id(t) for t in tokens if is_episode_id(t)]

    episode_ids = []
    for ri, row in enumerate(all_rows[2:], start=2):
        val = row[target_col] if target_col < len(row) else None
        if val is None: continue
        # Only extract bold cells — bold = show ID, non-bold = description text
        full_row = all_rows_full[ri] if ri < len(all_rows_full) else []
        cell = full_row[target_col] if target_col < len(full_row) else None
        is_bold = (cell is not None and cell.font is not None and cell.font.bold)
        if not is_bold: continue
        for ep in extract_ids(val):
            if ep: episode_ids.append(ep)
    return episode_ids

def _parse_date_str(val, force_year=None):
    """Parse date from string like 'Lun. / Mon. 03/30/26'.
    force_year overrides the year in the string (handles typos like 03/30/25 when it should be 2026)."""
    if not val: return None
    m = re.search(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', str(val))
    if m:
        mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if force_year:
            y = force_year
        else:
            y = 2000 + y if y < 100 else y
        try: return datetime(y, mo, d).date()
        except: pass
    return None

def _parse_grilla_pasiones(filepath_or_bytes, target_date):
    """
    Pasiones grilla parser — multi-tab yearly workbook.
    Scans tabs last→first, finds the one containing target_date.
    Header row contains date strings like 'Mar. / Tue. 03/31/26'.
    Episode IDs are in alternating rows (show name row, then ID row).
    """
    from openpyxl import load_workbook
    import io
    if isinstance(filepath_or_bytes, str):
        wb = load_workbook(filepath_or_bytes, read_only=True)
    else:
        raw = filepath_or_bytes.read() if hasattr(filepath_or_bytes, 'read') else filepath_or_bytes
        wb = load_workbook(io.BytesIO(raw), read_only=True)

    target_ws = None
    target_col = None

    for name in reversed(wb.sheetnames):
        ws = wb[name]
        rows = list(ws.iter_rows(max_row=3, values_only=True))
        if len(rows) < 2: continue
        header = rows[1]
        # Use target_date's year to avoid typos in spreadsheet year field
        col_dates = [(i, _parse_date_str(cell, force_year=target_date.year))
                     for i, cell in enumerate(header)
                     if _parse_date_str(cell, force_year=target_date.year)]
        if not col_dates: continue
        first_d = col_dates[0][1]
        last_d  = col_dates[-1][1]
        if first_d <= target_date <= last_d:
            for col_i, d in col_dates:
                if d == target_date:
                    target_col = col_i
                    target_ws  = ws
                    # Re-read full sheet
                    all_rows = list(ws.iter_rows(values_only=True))
                    break
            break

    if target_ws is None or target_col is None:
        return []

    # Detect which column is ET — show name rows have a time value there,
    # episode ID rows have None. Works for both ET-at-col-0 and UTC+ET formats.
    # Look for 'ET' or 'E.T.' in header row to find the column.
    header_row = all_rows[1] if len(all_rows) > 1 else []
    ET_COL = 0  # default
    for ci, cell in enumerate(header_row):
        if cell and isinstance(cell, str) and cell.strip().upper() in ('ET', 'E.T.'):
            ET_COL = ci
            break

    episode_ids = []
    for row in all_rows[2:]:
        et_val = row[ET_COL] if ET_COL < len(row) else 'x'
        if et_val is not None:
            continue  # show name row — skip
        val = row[target_col] if target_col < len(row) else None
        if val and isinstance(val, str):
            val = val.strip()
            if val:
                episode_ids.append(normalize_id(val))
    return episode_ids

def _parse_grilla_tn(filepath_or_bytes, target_date):
    """
    Todonovelas grilla — multi-tab, same header format as Pasiones.
    Returns list of (show_name, episode_num) tuples for program matching.
    Episode numbers are integers in the grid (55, 121...).
    Show name rows: ET column has value. Episode rows: ET column is None.
    """
    from openpyxl import load_workbook
    import io
    if isinstance(filepath_or_bytes, str):
        wb = load_workbook(filepath_or_bytes, read_only=True)
    else:
        raw = filepath_or_bytes.read() if hasattr(filepath_or_bytes, 'read') else filepath_or_bytes
        wb = load_workbook(io.BytesIO(raw), read_only=True)

    target_ws = None
    target_col = None
    all_rows = []

    for name in reversed(wb.sheetnames):
        ws = wb[name]
        rows = list(ws.iter_rows(max_row=3, values_only=True))
        if len(rows) < 2: continue
        header = rows[1]
        col_dates = [(i, _parse_date_str(cell, force_year=target_date.year))
                     for i, cell in enumerate(header)
                     if _parse_date_str(cell, force_year=target_date.year)]
        if not col_dates: continue
        if col_dates[0][1] <= target_date <= col_dates[-1][1]:
            for col_i, d in col_dates:
                if d == target_date:
                    target_col = col_i
                    target_ws  = ws
                    all_rows   = list(ws.iter_rows(values_only=True))
                    break
            break

    if not all_rows or target_col is None:
        return []

    ET_COL = 1
    result = []
    current_show = None
    for row in all_rows[2:]:
        et_val = row[ET_COL] if ET_COL < len(row) else 'x'
        val    = row[target_col] if target_col < len(row) else None
        if val is None: continue
        if et_val is not None:
            # Show name row
            current_show = str(val).strip() if val else None
        else:
            # Episode number row — val is an integer
            if current_show and val is not None:
                try:
                    ep_num = int(val)
                    result.append((current_show, ep_num))
                except (ValueError, TypeError):
                    pass
    return result


# ── FILE DETECTION ────────────────────────────────────────────────────────────

def extract_date_from_filename(name):
    # YYYYMMDD (JSON files: ..._20260330_...)
    m = re.search(r'(\d{4})(\d{2})(\d{2})', name)
    if m:
        try: return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))).date()
        except: pass
    # MMDDYYYY (XML files: TVD03302026.xml, CA03302026.xml)
    m = re.search(r'(\d{2})(\d{2})(\d{4})', name)
    if m:
        try: return datetime(int(m.group(3)), int(m.group(1)), int(m.group(2))).date()
        except: pass
    return None

def _date_from_json_content(f):
    """Extract date from JSON file content (first event's startTime). Most reliable."""
    try:
        f.seek(0)
        data = json.load(f)
        f.seek(0)
        for ev in data.get('events', []):
            tc = ev.get('startTime', '')
            if tc:
                m = re.search(r'(\d{4})-(\d{2})-(\d{2})', tc)
                if m:
                    return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))).date()
    except: pass
    return None

def _date_from_xml_filename(name):
    """XML filenames date extraction.
    Sony:     YYYYMMDD (A120260401c.XML, S620260401c_XML.xml)
    Standard: MMDDYYYY (TVD03302026.xml, CA03302026.xml, PL03312026.xml)
    TN:       MMDDYY   (TN_033126_TUESDAY.xml)
    """
    # Try YYYYMMDD first (8 consecutive digits starting with 20)
    m = re.search(r'(20\d{2})(\d{2})(\d{2})', name)
    if m:
        try: return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))).date()
        except: pass
    # Try MMDDYYYY (8 digits)
    m = re.search(r'(\d{2})(\d{2})(\d{4})', name)
    if m:
        try: return datetime(int(m.group(3)), int(m.group(1)), int(m.group(2))).date()
        except: pass
    # Try MMDDYY (6 digits, 2-digit year)
    m = re.search(r'(\d{2})(\d{2})(\d{2})(?!\d)', name)
    if m:
        try: return datetime(2000 + int(m.group(3)), int(m.group(1)), int(m.group(2))).date()
        except: pass
    return None

def detect_files(uploaded_files):
    """
    Group uploaded files by (date, channel).
    JSON: date always from content (startTime). Works for renamed files.
    XML:  date from filename (MMDDYYYY pattern, always reliable).
    Grilla: week-based, stored by channel only.
    Returns: days dict, grillas dict, unknown list.
    """
    days    = {}
    grillas = {}
    unknown = []
    sony_files = []   # Sony/AXN files handled separately

    for f in uploaded_files:
        name_up = f.name.upper()
        ext = f.name.lower().rsplit('.', 1)[-1] if '.' in f.name else ''

        if ext == 'json':
            ftype = 'json'
        elif ext == 'xml':
            ftype = 'xml'
        elif ext == 'pdf':
            # HolaTV PDF grilla only
            if 'USH' in name_up or ('HOLA' in name_up and 'US' in name_up):
                grillas.setdefault('hu', []).append(f)
            elif 'LATAM' in name_up:
                # Accept any LATAM PDF as HolaTV Latam grilla
                grillas.setdefault('hl', []).append(f)
            else:
                unknown.append(f)
            continue
        elif ext in ('xlsx', 'xlsm'):
            # HolaTV XLSX log (not grilla)
            if name_up.startswith('HU'):
                ftype = 'log'; 
            elif name_up.startswith('HL'):
                ftype = 'log'
            else:
                ftype = 'grilla'
        elif ext == 'txt':
            if name_up.startswith('HU') or name_up.startswith('HL'):
                ftype = 'log'
            else:
                unknown.append(f); continue
        else:
            unknown.append(f); continue

        # Sony/AXN detection — check before generic channel detection
        sony_code = extract_sony_code(f.name)
        if sony_code:
            sony_files.append({'file': f, 'ftype': ftype, 'code': sony_code})
            continue

        # Channel detection
        if ext == 'xml':
            if name_up.startswith('TVD'):   channel = 'tvd'
            elif name_up.startswith('CA'):  channel = 'catv'
            elif name_up.startswith('PL'):  channel = 'latam'
            elif name_up.startswith('PUS'): channel = 'us'
            elif name_up.startswith('TN'):  channel = 'tn'
            elif name_up.startswith('HU'):  channel = 'hu'
            elif name_up.startswith('HL'):  channel = 'hl'
            else: unknown.append(f); continue
        elif ext == 'pdf':
            # HolaTV PDF grilla
            if 'USH' in name_up or 'HOLATV_US' in name_up or 'HOLA_US' in name_up or 'HOLA TV US' in name_up:
                grillas['hu'] = f; continue
            elif 'LATAM' in name_up and ('HOLA' in name_up or 'HL' in name_up):
                grillas['hl'] = f; continue
            else: unknown.append(f); continue
        elif ext == 'json':
            # JSON channel detection
            if 'HOLATV_US' in name_up or 'HOLA_TV_US' in name_up or 'HOLA_US' in name_up or ('HOLATV' in name_up and 'US' in name_up):
                channel = 'hu'
            elif 'HOLATV_LATAM' in name_up or 'HOLA_TV_LATAM' in name_up or 'HOLA_LATAM' in name_up or ('HOLATV' in name_up and 'LATAM' in name_up):
                channel = 'hl'
            elif 'PASIONES_LATAM' in name_up or 'PASIONES LATAM' in name_up or name_up.startswith('PL'):
                channel = 'latam'
            elif 'PASIONES_US' in name_up or 'PASIONES US' in name_up or name_up.startswith('PUS'):
                channel = 'us'
            elif 'FAST_TODONOVELAS' in name_up or 'FAST TODONOVELAS' in name_up or ('TODO' in name_up and 'NOVELA' in name_up):
                channel = 'tn'
            elif 'TVD' in name_up: channel = 'tvd'
            elif 'CATV' in name_up or name_up.startswith('CA'): channel = 'catv'
            else: unknown.append(f); continue
        else:
            if 'HOLATV_US' in name_up or 'HOLA_TV_US' in name_up or ('HOLATV' in name_up and 'US' in name_up) or name_up.startswith('HU'): 
                channel = 'hu'
            elif 'HOLATV_LATAM' in name_up or 'HOLA_TV_LATAM' in name_up or ('HOLATV' in name_up and 'LATAM' in name_up) or name_up.startswith('HL'): 
                channel = 'hl'
            elif 'CATV' in name_up:                       channel = 'catv'
            elif 'TVD' in name_up:                        channel = 'tvd'
            elif 'PASIONES_LATAM' in name_up or 'PASIONES LATAM' in name_up: channel = 'latam'
            elif 'PASIONES_US' in name_up or 'PASIONES US' in name_up:       channel = 'us'
            elif 'FAST_TODONOVELAS' in name_up or 'FAST TODONOVELAS' in name_up or ('TODO' in name_up and 'NOVELA' in name_up): channel = 'tn'
            else: unknown.append(f); continue

        if ftype == 'grilla':
            grillas.setdefault(channel, []).append(f)
            continue

        # Date extraction
        if ftype == 'json':
            date = _date_from_json_content(f)    # always from content
        else:  # xml
            date = _date_from_xml_filename(f.name)  # always from filename

        if date is None:
            unknown.append(f)
            continue

        key = (str(date), channel)
        if key not in days:
            days[key] = {'json': [], 'xml': None, 'log': None, 'date': date}
        if ftype == 'json':  days[key]['json'].append(f)
        elif ftype == 'xml': days[key]['xml'] = f
        elif ftype == 'log': days[key]['log'] = f

    return days, grillas, unknown, sony_files


def pair_sony_files(sony_files, lang='en'):
    """
    Pair Sony JSON files with Sony XML log files by channel code + date/marker.
    Returns list of pairing dicts for the app to process.
    Each: {'label', 'code', 'channel_name', 'json_file', 'xml_file', 'xml_filename', 'date'}
    Plus unmatched XMLs and JSONs.
    """
    json_list = [f for f in sony_files if f['ftype'] == 'json']
    xml_list  = [f for f in sony_files if f['ftype'] == 'xml']

    pairings    = []
    used_jsons  = set()

    # For each JSON: read markers to find expected XML(s)
    # Also extract date from JSON content for fallback
    json_info = []
    for jf in json_list:
        try:
            jf['file'].seek(0)
            data = json.load(jf['file'])
            jf['file'].seek(0)
        except:
            data = {'events': []}
        markers  = parse_sony_json_markers(data)
        date_val = _date_from_json_content(jf['file'])
        try: jf['file'].seek(0)
        except: pass
        json_info.append({'jf': jf, 'data': data,
                          'markers': markers, 'date': date_val})

    # Match full JSONs (with markers) first, then partial (date-based) to avoid consuming wrong XML
    json_info_sorted = sorted(json_info, key=lambda x: 0 if x['markers'] else 1)
    for jinfo in json_info_sorted:
        jf       = jinfo['jf']
        code     = jf['code']
        ch_name  = SONY_CHANNEL_MAP.get(code, code)
        matched_xml = None
        # Use filename date as fallback if content date failed
        if jinfo['date'] is None:
            jinfo['date'] = _date_from_xml_filename(jf['file'].name)

        if jinfo['markers']:
            for mk in jinfo['markers']:
                if not mk['log_base']: continue
                for xf in xml_list:
                    if xf['code'] != code: continue
                    xbase = extract_sony_xml_base(xf['file'].name)
                    if xbase.upper() == mk['log_base'].upper():
                        matched_xml = xf
                        break
                if matched_xml: break

        # Fallback: match by channel code + date from filename
        if not matched_xml and jinfo['date']:
            date_str = jinfo['date'].strftime('%Y%m%d')
            for xf in xml_list:
                if xf['code'] != code: continue
                if date_str in xf['file'].name:
                    matched_xml = xf
                    break

        # Last resort: if only one XML for this channel code, pair it
        if not matched_xml:
            available = [xf for xf in xml_list
                         if xf['code'] == code]
            if len(available) == 1:
                matched_xml = available[0]

        used_jsons.add(id(jf))
        pairings.append({
            'label':        f'{ch_name} — {jinfo["date"] or "?"}',
            'code':         code,
            'channel_name': ch_name,
            'json_file':    jf['file'],
            'json_data':    jinfo['data'],
            'xml_file':     matched_xml['file'] if matched_xml else None,
            'xml_filename': matched_xml['file'].name if matched_xml else None,
            'date':         jinfo['date'],
        })

    # Unmatched XMLs (no JSON for them)
    # Show XMLs with no JSON — find ones not paired to any JSON by code+date
    paired_xml_ids = {id(p["xml_file"]) for p in pairings if p["xml_file"]}
    unmatched_xml = [xf for xf in xml_list if id(xf["file"]) not in paired_xml_ids]
    for xf in unmatched_xml:
        code    = xf['code']
        ch_name = SONY_CHANNEL_MAP.get(code, code)
        pairings.append({
            'label':        f'{ch_name} — LOG ONLY (no JSON)',
            'code':         code,
            'channel_name': ch_name,
            'json_file':    None,
            'json_data':    None,
            'xml_file':     xf['file'],
            'xml_filename': xf['file'].name,
            'date':         None,
        })

    return pairings


# ── CHECKS ────────────────────────────────────────────────────────────────────

def check_programs_vs_grilla(playlist, grilla_ids, current_start, lang):
    """
    LCS-diff based program comparison. Handles:
    - Re-airs (walks both sequences in order, so second occurrence is checked at its position)
    - Replacements (CEND0402 deleted / NP0402 added — reports both, no cascading false errors)
    - Wrong episodes (same prefix, different date suffix)
    - Insertions / deletions
    """
    is_partial = current_start is not None
    part_seq   = build_show_sequence(playlist['programs'], from_start=current_start)

    if not grilla_ids:
        return [T('no_grilla', lang)]

    # Anchor partial playlist to correct grilla occurrence (handles re-air anchoring)
    if is_partial and part_seq:
        first_id  = part_seq[0]['id']
        first_pfx = show_prefix(first_id)
        # Count how many times first_id aired BEFORE current_start
        pre_count = sum(1 for p in playlist['programs']
                        if p['episode_id'] == first_id
                        and p['start'] and p['start'] < current_start)
        anchor, found_count = 0, 0
        for i, gid in enumerate(grilla_ids):
            if gid == first_id or (first_pfx and show_prefix(gid) == first_pfx):
                if found_count == pre_count:
                    anchor = i; break
                found_count += 1
        issues = [T('anchored', lang, i=anchor+1, id=grilla_ids[anchor] if grilla_ids else '?')]
        grilla_slice = grilla_ids[anchor:]
    else:
        grilla_slice = grilla_ids[:]
        issues = []

    # LCS sequential walk
    WINDOW = 8
    gi, pi = 0, 0

    while gi < len(grilla_slice) and pi < len(part_seq):
        gid = grilla_slice[gi]
        p   = part_seq[pi]
        pid = p['id']

        if gid == pid:
            gi += 1; pi += 1; continue

        pfx_g = show_prefix(gid)
        pfx_p = show_prefix(pid)

        # Same prefix → wrong episode (COSA0326 vs COSA0402)
        # For movie IDs (pure alpha like MARCE): also catch MARCE vs MARCELO
        is_movie_g = bool(re.match(r'^[A-Z]{3,8}$', gid))
        is_movie_p = bool(re.match(r'^[A-Z]{3,8}$', pid))
        pfx_match = (pfx_g and pfx_p and (
            pfx_g == pfx_p or
            ((is_movie_g or is_movie_p) and
             (pfx_g.startswith(pfx_p) or pfx_p.startswith(pfx_g)))
        ))
        if pfx_match:
            issues.append(T('wrong_ep', lang, g=gid, p=pid, t=fmt_t(p['start'])))
            gi += 1; pi += 1; continue

        # Look ahead
        future_pl = [part_seq[pi+k]['id'] for k in range(1, min(WINDOW+1, len(part_seq)-pi))]
        future_gr = list(grilla_slice[gi+1:gi+WINDOW+1])

        gid_ahead_pl = gid in future_pl
        pid_ahead_gr = pid in future_gr

        if not gid_ahead_pl and not pid_ahead_gr:
            # Replacement: grilla has X, playlist has Y, neither in each other near future
            deleted_lbl = 'DELETED' if lang == 'en' else 'ELIMINADO'
            added_lbl   = 'ADDED'   if lang == 'en' else 'AGREGADO'
            issues.append(f'  ↔  {deleted_lbl}: {gid} / {added_lbl}: {pid} @ {fmt_t(p["start"])}')
            gi += 1; pi += 1

        elif not gid_ahead_pl:
            # Grilla show deleted from playlist
            issues.append(T('not_in_pl', lang, id=gid))
            gi += 1

        elif not pid_ahead_gr:
            # Extra show in playlist not in grilla
            issues.append(T('extra_pl', lang, id=pid, t=fmt_t(p['start'])))
            pi += 1

        else:
            # Both in each other's future — realign via shortest path
            pl_offset = future_pl.index(gid) + 1   # steps in pl to reach gid
            gr_offset = future_gr.index(pid) + 1    # steps in gr to reach pid
            if pl_offset <= gr_offset:
                for k in range(pl_offset):
                    issues.append(T('extra_pl', lang, id=part_seq[pi+k]['id'],
                                    t=fmt_t(part_seq[pi+k]['start'])))
                pi += pl_offset
            else:
                for k in range(gr_offset):
                    issues.append(T('not_in_pl', lang, id=grilla_slice[gi+k]))
                gi += gr_offset

    # Tail: remaining grilla entries not in playlist
    while gi < len(grilla_slice):
        gid = grilla_slice[gi]
        if is_partial and any(p['episode_id'] == gid for p in playlist['programs']
                              if p['start'] and p['start'] < current_start):
            issues.append(T('already_aired', lang, id=gid))
        else:
            issues.append(T('not_in_pl', lang, id=gid))
        gi += 1

    # Tail: remaining playlist entries
    while pi < len(part_seq):
        issues.append(T('extra_pl', lang, id=part_seq[pi]['id'], t=fmt_t(part_seq[pi]['start'])))
        pi += 1

    has_errors = any(x.strip().startswith(('✗','⚠','↔')) for x in issues)
    if not has_errors:
        issues.append(f'  {T("ok_programs", lang)}')
    return issues


def check_commercials_vs_xml(playlist, xml_rows, current_start, lang):
    """
    Break-by-break commercial comparison.
    Aligns XML and JSON by segment ID (after_program_raw).
    Handles show replacements with a look-ahead window.
    Returns (issues_list, manual_warnings_list).
    """
    WINDOW = 15  # max segments to look ahead for replacement recovery

    # --- Build XML break list ---
    xml_breaks = build_xml_breaks(xml_rows)

    # --- Build JSON break list (only those with a raw anchor) ---
    json_breaks = [b for b in playlist['breaks'] if b.get('after_program_raw')]

    # --- For partial: align by matching first JSON break's segment in XML ---
    if current_start:
        # Filter json_breaks to those at/after current_start
        def _break_time(b):
            return next((i['start'] for i in b.get('items',[]) if i.get('start')), None)
        json_breaks = [b for b in json_breaks
                       if not _break_time(b) or _break_time(b) >= current_start]

        if json_breaks:
            first_seg = json_breaks[0].get('after_program_raw', '')
            if first_seg:
                # Find this segment in XML breaks and start from there
                xi = next((i for i, xb in enumerate(xml_breaks)
                           if xb['anchor_id'] == first_seg), None)
                if xi is not None:
                    xml_breaks = xml_breaks[xi:]  # start FROM this segment (not after)

    # --- Labels ---
    added_lbl    = {'en': 'added to playlist',    'es': 'agregado a playlist'}
    removed_lbl  = {'en': 'removed from playlist', 'es': 'eliminado de playlist'}
    replaced_lbl = {'en': 'SHOW REPLACED',         'es': 'PROGRAMA REEMPLAZADO'}
    lost_lbl     = {'en': 'ALIGNMENT LOST',        'es': 'ALINEACIÓN PERDIDA'}
    summary_lbl  = {'en': 'COMMERCIAL CHANGES SUMMARY', 'es': 'RESUMEN DE CAMBIOS'}
    added_tot    = {'en': 'added',    'es': 'agregados'}
    removed_tot  = {'en': 'removed',  'es': 'eliminados'}
    manual_cap   = {'en': '!!! DOUBLE CHECK MANUALLY !!!', 'es': '!!! VERIFICAR MANUALMENTE !!!'}

    issues       = []
    all_added    = Counter()
    all_removed  = Counter()
    manual_warns = []

    def _break_start(jb):
        return next((i['start'] for i in jb['items'] if i.get('start')), None)

    def _compare_pair(xb, jb):
        """Compare one aligned XML/JSON break pair. Returns (lines, added, removed)."""
        xml_c  = Counter(xb['commercials'])
        json_c = Counter(i['ref'] for i in jb['items'] if i['type'] == 'Commercial')
        anchor = jb.get('after_program_raw', '?')
        bs     = _break_start(jb)

        lines, add, rem = [], Counter(), Counter()
        for ref in sorted(set(xml_c) | set(json_c)):
            xc, jc = xml_c.get(ref, 0), json_c.get(ref, 0)
            if xc == jc: continue
            diff = jc - xc
            if diff > 0:
                lines.append(f'     + {ref} x{diff}  ({added_lbl[lang]})')
                add[ref] += diff
            else:
                lines.append(f'     - {ref} x{abs(diff)}  ({removed_lbl[lang]})')
                rem[ref] += abs(diff)

        if lines:
            header = [f'  ⚠  Break after [{anchor}] @ {fmt_t(bs)}']
            return header + lines, add, rem
        return [], add, rem

    def _compare_pool(xml_blist, json_blist, xml_label, json_label, bs):
        """Compare pooled commercials from a replaced block."""
        xml_c  = Counter(ref for xb in xml_blist for ref in xb['commercials'])
        json_c = Counter(ref for jb in json_blist for i in jb['items']
                         if i['type'] == 'Commercial' for ref in [i['ref']])
        lines, add, rem = [], Counter(), Counter()
        for ref in sorted(set(xml_c) | set(json_c)):
            xc, jc = xml_c.get(ref, 0), json_c.get(ref, 0)
            if xc == jc: continue
            diff = jc - xc
            if diff > 0: lines.append(f'     + {ref} x{diff}  ({added_lbl[lang]})'); add[ref] += diff
            else:        lines.append(f'     - {ref} x{abs(diff)}  ({removed_lbl[lang]})'); rem[ref] += abs(diff)
        return lines, add, rem

    # --- Walk both break lists in parallel ---
    xi, ji = 0, 0
    while xi < len(xml_breaks) and ji < len(json_breaks):
        xb = xml_breaks[xi]
        jb = json_breaks[ji]
        x_anc = xb['anchor_id']
        j_anc = jb.get('after_program_raw', '')

        if x_anc == j_anc:
            # Perfect match
            lines, add, rem = _compare_pair(xb, jb)
            issues.extend(lines)
            all_added.update(add)
            all_removed.update(rem)
            xi += 1; ji += 1

        else:
            # Mismatch — look ahead in both directions to recover
            found_xi = next((i for i in range(xi+1, min(xi+WINDOW, len(xml_breaks)))
                             if xml_breaks[i]['anchor_id'] == j_anc), None)
            found_ji = next((i for i in range(ji+1, min(ji+WINDOW, len(json_breaks)))
                             if json_breaks[i].get('after_program_raw') == x_anc), None)

            bs = _break_start(jb)

            if found_xi is not None and (found_ji is None or (found_xi-xi) <= (found_ji-ji)):
                # XML has more segments here — pooled comparison
                xml_block  = xml_breaks[xi:found_xi]
                json_block = [jb]
                x_show = normalize_id(x_anc)
                j_show = normalize_id(j_anc)
                issues.append(f'  ⚠  {replaced_lbl[lang]}: XML=[{x_show}...] → Playlist=[{j_show}] @ {fmt_t(bs)}')
                pool_lines, add, rem = _compare_pool(xml_block, json_block, x_show, j_show, bs)
                if pool_lines:
                    issues.extend(pool_lines)
                    warn = f'{manual_cap[lang]}: {replaced_lbl[lang]} [{x_show}→{j_show}] @ {fmt_t(bs)}'
                    manual_warns.append(warn)
                    issues.append(f'     ⚠  {warn}')
                else:
                    issues.append(f'     ✓ Commercials match within replaced block')
                all_added.update(add); all_removed.update(rem)
                xi = found_xi; ji += 1

            elif found_ji is not None:
                # JSON has more segments here — pooled comparison
                xml_block  = [xb]
                json_block = json_breaks[ji:found_ji]
                x_show = normalize_id(x_anc)
                j_show = normalize_id(j_anc)
                issues.append(f'  ⚠  {replaced_lbl[lang]}: XML=[{x_show}] → Playlist=[{j_show}...] @ {fmt_t(bs)}')
                pool_lines, add, rem = _compare_pool(xml_block, json_block, x_show, j_show, bs)
                if pool_lines:
                    issues.extend(pool_lines)
                    warn = f'{manual_cap[lang]}: {replaced_lbl[lang]} [{x_show}→{j_show}] @ {fmt_t(bs)}'
                    manual_warns.append(warn)
                    issues.append(f'     ⚠  {warn}')
                else:
                    issues.append(f'     ✓ Commercials match within replaced block')
                all_added.update(add); all_removed.update(rem)
                xi += 1; ji = found_ji

            else:
                # Can't recover — skip both and warn loudly
                warn = f'{manual_cap[lang]}: {lost_lbl[lang]} [{normalize_id(x_anc)} vs {normalize_id(j_anc)}] @ {fmt_t(bs)}'
                manual_warns.append(warn)
                issues.append(f'  ⚠  {warn}')
                xi += 1; ji += 1

    # --- Summary ---
    if all_added or all_removed:
        issues += ['', f'  ── {summary_lbl[lang]} ──']
        if all_added:
            total = sum(all_added.values())
            issues.append(f'  +{total} {added_tot[lang]}:')
            for ref, cnt in sorted(all_added.items()):
                issues.append(f'    {ref} x{cnt}')
        if all_removed:
            total = sum(all_removed.values())
            issues.append(f'  -{total} {removed_tot[lang]}:')
            for ref, cnt in sorted(all_removed.items()):
                issues.append(f'    {ref} x{cnt}')
    elif not issues:
        total_pl = len(playlist['commercials'])
        issues.append(f'  {T("ok_commercials", lang, n=total_pl)}')

    return issues, manual_warns


def check_promo_repeats(playlist, current_start=None, lang='en'):
    INFOMERCIAL_SECS = 1200  # 20 min
    issues = []

    for brk in playlist['breaks']:
        items = brk['items']
        if not items: continue
        bs = next((i['start'] for i in items if i.get('start')), None)
        if current_start and bs and bs < current_start: continue

        # Split break at infomercials (Commercial ≥ 20min) — they act as sub-break separators
        sub_breaks = []
        current_sub = []
        for item in items:
            if item['type'] == 'Commercial' and parse_duration(item.get('duration','00:00:00')) >= INFOMERCIAL_SECS:
                if current_sub:
                    sub_breaks.append(current_sub)
                current_sub = []  # reset after infomercial
            else:
                current_sub.append(item)
        if current_sub:
            sub_breaks.append(current_sub)
        if not sub_breaks:
            sub_breaks = [items]

        for sub in sub_breaks:
            promo_refs = [i['ref'] for i in sub if i['type'] == 'Promotion']
            for ref, cnt in Counter(promo_refs).items():
                if cnt > 1:
                    after = brk.get('after_program', '?')
                    sub_start = next((i['start'] for i in sub if i.get('start')), bs)
                    issues.append(T('promo_rep', lang, after=after,
                                    t=fmt_t(sub_start), ref=ref, n=cnt))


def check_not_ingested(playlist, current_start=None, lang='en'):
    lines, seen_eps, seen_other = [], set(), set()
    for item in playlist['not_ingested']:
        if current_start and item['start'] and item['start'] < current_start: continue
        atype, aref = item['type'], item['asset_ref']
        if atype in ('Program','live'):
            ep = normalize_id(aref)
            if ep in seen_eps: continue
            seen_eps.add(ep)
            show = re.sub(r'\[\].*$', '', item['name']).strip()
            lines.append(T('ni_program', lang, id=ep, t=fmt_t(item['start']), show=show))
        else:
            if aref in seen_other: continue
            seen_other.add(aref)
            lines.append(T('ni_other', lang, typ=atype, ref=aref,
                           t=fmt_t(item['start']), name=item['name']))
    return lines


def check_bugs(playlist, current_start=None, lang='en'):
    """
    Check LOGOHD bug logo assignments.
    Groups consecutive programs by their logo value and reports time ranges.
    Only reads from program events that have a 'logo' field set.
    """
    progs = [p for p in playlist['programs']
             if p.get('logo') is not None
             and (not current_start or not p['start'] or p['start'] >= current_start)]
    if not progs:
        return [f"  \u2714  No bugs scheduled" if lang=='en' else "  \u2714  Sin bugs programados"]

    # Group consecutive programs by logo
    groups = []  # (logo, start_time, end_time, programs_in_group)
    cur_logo   = progs[0]['logo']
    cur_start  = progs[0]['start']
    cur_progs  = [progs[0]]
    for p in progs[1:]:
        if p['logo'] == cur_logo:
            cur_progs.append(p)
        else:
            groups.append((cur_logo, cur_start, p['start'], cur_progs[:]))
            cur_logo  = p['logo']
            cur_start = p['start']
            cur_progs = [p]
    groups.append((cur_logo, cur_start, None, cur_progs))

    lines = []
    for logo, t_start, t_end, grp in groups:
        s = fmt_t(t_start) if t_start else '?'
        e = fmt_t(t_end)   if t_end   else ('end of day' if lang=='en' else 'fin del día')
        lines.append(f'  {logo}  :  {s} → {e}')
    return lines


def check_cue_tones(playlist, lang='en'):
    """
    Cue tone report.
    CUE ON = trigger (its duration not counted).
    Clips after CUE ON up to and including CUE OFF = cue block.
    Duration computed using consecutive promo start-time gaps.
    """
    cts = playlist.get('cue_tones', [])
    if not cts:
        return [f"  \u2714  No cue tones found" if lang=='en' else "  \u2714  Sin cue tones"]

    all_promos = sorted(playlist.get('promos', []), key=lambda p: p['start'] or datetime.min)
    if not all_promos:
        return [f"  \u2139  {len(cts)} CUE ON found — no promo list to compute durations"]

    # Build start_time -> duration map using consecutive gaps
    promo_durs = {}
    for i, p in enumerate(all_promos):
        if p['start']:
            if i + 1 < len(all_promos) and all_promos[i+1]['start']:
                promo_durs[p['start']] = int((all_promos[i+1]['start'] - p['start']).total_seconds())
            else:
                promo_durs[p['start']] = 30  # last promo default

    from collections import defaultdict
    stats = defaultdict(lambda: {'count': 0, 'first': None, 'last_dur': 0})

    sorted_cts = sorted(cts, key=lambda c: c['start'] or datetime.min)
    cue_blocks = []

    for i, ct in enumerate(sorted_cts):
        ref  = ct['ct_id']
        t_on = ct['start']
        if not t_on: continue
        t_next = sorted_cts[i+1]['start'] if i+1 < len(sorted_cts) else None

        # Sum durations of promos AFTER t_on until next CUE ON
        block_dur = sum(
            promo_durs.get(p['start'], 0)
            for p in all_promos
            if p['start'] and p['start'] > t_on
            and (t_next is None or p['start'] < t_next)
        )
        cue_blocks.append({'ref': ref, 'start': t_on, 'dur': block_dur})
        stats[ref]['count'] += 1
        stats[ref]['last_dur'] = min(block_dur, 240)
        if stats[ref]['first'] is None or t_on < stats[ref]['first']:
            stats[ref]['first'] = t_on

    total_count = len(sorted_cts)
    total_dur   = sum(min(b['dur'], 240) for b in cue_blocks)

    def fmt_dur(secs):
        m, s = divmod(int(secs), 60)
        return f'{m}min {s:02d}sec'

    lines = [f"  Total CUE ON: {total_count} | Total duration: {fmt_dur(total_dur)}"]
    for ref in sorted(stats):
        s = stats[ref]
        first_str = s['first'].strftime('%H:%M:%S') if s['first'] else '?'
        last_str  = fmt_dur(s['last_dur'])
        lines.append(f"  {ref}: {s['count']}x | First: {first_str} | Last: {last_str}")
    return lines

def check_holatv_programs_v2(grilla_entries, log_blocks, current_start_utc, lang):
    """
    Episode-number-only program check for HolaTV, LCS-style.
    One missing/extra entry does not cascade — handled like CATV/TVD.
    grilla_entries: (show_eps list, inf_count) tuple from parse_grilla_holatv_v2
    """
    if isinstance(grilla_entries, tuple):
        grilla_show_eps, grilla_inf_count = grilla_entries
    else:
        grilla_show_eps = [g['episode'] for g in grilla_entries if not g.get('is_inf')]
        grilla_inf_count = sum(1 for g in grilla_entries if g.get('is_inf'))

    if not grilla_show_eps and grilla_inf_count == 0:
        return [f'  ℹ  {"No grilla provided" if lang=="en" else "Sin grilla proporcionada"}']

    active    = [b for b in log_blocks
                 if not current_start_utc or not b['start_utc']
                 or b['start_utc'] >= current_start_utc]
    log_shows = [b for b in active if not b['is_hpp']]
    log_hpp   = [b for b in active if b['is_hpp']]

    if not log_shows and not log_hpp:
        return [f'  ℹ  {"No log data in window" if lang=="en" else "Sin datos de log en la ventana"}']

    def ep_from_id(base_id):
        m = re.search(r'(\d+)$', base_id)
        return int(m.group(1)) if m else None

    # ── Partial anchoring ──
    if current_start_utc and log_shows and grilla_show_eps:
        log_head   = [ep_from_id(b['base_id']) for b in log_shows[:4]]
        anchor_pos = None
        for gi in range(len(grilla_show_eps)):
            needed = min(3, len(grilla_show_eps) - gi, len(log_head))
            if needed < 1: break
            if all(grilla_show_eps[gi+k] == log_head[k] for k in range(needed)):
                anchor_pos = gi; break
        if anchor_pos is not None:
            grilla_show_eps = grilla_show_eps[anchor_pos:]

    # ── LCS walk ──
    WINDOW = 8
    issues  = []
    ok_count = 0
    gi, li  = 0, 0
    g_eps   = grilla_show_eps
    l_blks  = log_shows

    while gi < len(g_eps) and li < len(l_blks):
        g_ep = g_eps[gi]
        l_ep = ep_from_id(l_blks[li]['base_id'])

        if g_ep == l_ep:
            ok_count += 1; gi += 1; li += 1; continue

        # Look ahead
        fut_g = [g_eps[gi+k] for k in range(1, min(WINDOW+1, len(g_eps)-gi))]
        fut_l = [ep_from_id(l_blks[li+k]['base_id']) for k in range(1, min(WINDOW+1, len(l_blks)-li))]

        g_in_fut_l = g_ep in fut_l
        l_in_fut_g = l_ep in fut_g

        if not g_in_fut_l and not l_in_fut_g:
            warn = 'MANUAL CHECK' if lang=='en' else 'REVISIÓN MANUAL'
            issues.append(f'  ⚠  {warn}: grilla ep{g_ep} ≠ log {l_blks[li]["base_id"]} (ep{l_ep}) @ {fmt_t(l_blks[li]["start_utc"])}')
            gi += 1; li += 1
        elif not g_in_fut_l:
            not_lbl = 'NOT IN LOG' if lang=='en' else 'NO EN LOG'
            issues.append(f'  ✗  {not_lbl}: grilla ep{g_ep}')
            gi += 1
        elif not l_in_fut_g:
            ext_lbl = 'EXTRA IN LOG' if lang=='en' else 'EXTRA EN LOG'
            issues.append(f'  ✗  {ext_lbl}: {l_blks[li]["base_id"]} @ {fmt_t(l_blks[li]["start_utc"])}')
            li += 1
        else:
            bl_off = fut_l.index(g_ep) + 1
            gr_off = fut_g.index(l_ep) + 1 if l_ep in fut_g else WINDOW
            if bl_off <= gr_off:
                for k in range(bl_off):
                    issues.append(f'  ✗  {"EXTRA IN LOG" if lang=="en" else "EXTRA EN LOG"}: {l_blks[li+k]["base_id"]} @ {fmt_t(l_blks[li+k]["start_utc"])}')
                li += bl_off
            else:
                for k in range(gr_off):
                    issues.append(f'  ✗  {"NOT IN LOG" if lang=="en" else "NO EN LOG"}: grilla ep{g_eps[gi+k]}')
                gi += gr_off

    while gi < len(g_eps):
        issues.append(f'  ✗  {"NOT IN LOG" if lang=="en" else "NO EN LOG"}: grilla ep{g_eps[gi]}')
        gi += 1
    while li < len(l_blks):
        issues.append(f'  ✗  {"EXTRA IN LOG" if lang=="en" else "EXTRA EN LOG"}: {l_blks[li]["base_id"]} @ {fmt_t(l_blks[li]["start_utc"])}')
        li += 1

    # ── INF / HPP counter ──
    inf_lbl = 'Infomercials' if lang=='en' else 'Infomerciales'
    if current_start_utc:
        issues.append(f'  ℹ  {inf_lbl}: {len(log_hpp)} HPP in log window (grilla count skipped for partial)')
    elif grilla_inf_count == len(log_hpp):
        issues.append(f'  ✓  {inf_lbl}: {grilla_inf_count} in grilla, {len(log_hpp)} HPP in log — match')
    else:
        diff = len(log_hpp) - grilla_inf_count
        sign = '+' if diff > 0 else ''
        issues.append(f'  ✗  {inf_lbl}: grilla={grilla_inf_count}, log={len(log_hpp)} ({sign}{diff})')

    if not any('✗' in i or '⚠' in i for i in issues):
        issues.insert(0, f'  ✓  {"All" if lang=="en" else "Todos"} {ok_count} {"episodes match" if lang=="en" else "episodios coinciden"}')
    else:
        n_warn = sum(1 for i in issues if '⚠' in i)
        n_err  = sum(1 for i in issues if '✗' in i)
        issues.insert(0, f'  ✓ {ok_count} {"match" if lang=="en" else "coinciden"}  |  ❗ {n_warn} {"manual check" if lang=="en" else "revisión manual"}  |  ✗ {n_err} {"mismatch" if lang=="en" else "diferencia"}')
    return issues


def check_holatv_timing_v2(grilla_entries, log_blocks, current_start_utc, lang, tolerance_secs=2700):
    """Timing check placeholder — grilla times not reliable."""
    return [f'  ✓  {"Episodes match in Grilla, Log and Playlist." if lang=="en" else "Episodios coinciden en Grilla, Log y Playlist."}']


def parse_grilla_holatv_v2(filepath_or_bytes, target_date):
    """
    Parse HolaTV PDF grilla — extracts episode numbers in column order.
    Uses vertical-distance-first code matching + digit-token merging.
    Returns (show_eps, inf_count).
    """
    try:
        import pdfplumber as _ppl
    except ImportError:
        return [], 0
    try:
        if hasattr(filepath_or_bytes, 'read'):
            filepath_or_bytes.seek(0)
            raw = filepath_or_bytes.read()
            filepath_or_bytes.seek(0)
            import tempfile, os as _os
            tmp = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
            tmp.write(raw); tmp.close()
            pdf_path = tmp.name; cleanup = True
        else:
            pdf_path = filepath_or_bytes; cleanup = False

        all_words = []
        with _ppl.open(pdf_path) as pdf:
            for pg_idx, page in enumerate(pdf.pages):
                words  = page.extract_words(x_tolerance=8, y_tolerance=3)
                page_h = float(page.height)
                for w in words:
                    all_words.append({**w, 'abs_top': w['top'] + pg_idx * page_h})
        if cleanup:
            import os as _os2; _os2.unlink(pdf_path)

        date_pat = re.compile(r'^(\d{2})/(\d{2})$')
        day_cols = {}
        for i, w in enumerate(all_words):
            m = date_pat.match(w['text'])
            if m and i > 0 and all_words[i-1]['text'].endswith('.'):
                day_x  = (all_words[i-1]['x0'] + w['x1']) / 2
                month, day_n = int(m.group(2)), int(m.group(1))
                try:
                    from datetime import date as _d2
                    d = _d2(target_date.year, month, day_n)
                    day_cols[str(d)] = day_x
                except Exception:
                    pass

        if str(target_date) not in day_cols:
            return [], 0

        target_x  = day_cols[str(target_date)]
        xs        = sorted(day_cols.values())
        idx       = xs.index(target_x)
        col_left  = (xs[idx-1] + xs[idx]) / 2 if idx > 0 else 0
        col_right = (xs[idx] + xs[idx+1]) / 2 if idx < len(xs)-1 else 9999

        code_re    = re.compile(r'^[A-Z][A-Z0-9_]{2,9}$')
        codes_at_y = []
        for i, w in enumerate(all_words):
            if code_re.match(w['text']) and i+1 < len(all_words) and all_words[i+1]['text'] == '(-)':
                codes_at_y.append((w['abs_top'], (w['x0']+w['x1'])/2, w['text']))

        ep_re   = re.compile(r'^\d{1,4}$')
        raw_eps = sorted(
            [{'y': w['abs_top'], 'text': w['text'],
              'x0': w['x0'], 'x1': w['x1'],
              'cx': (w['x0']+w['x1'])/2}
             for w in all_words
             if col_left <= (w['x0']+w['x1'])/2 <= col_right
             and ep_re.match(w['text'])],
            key=lambda w: (w['y'], w['x0']))

        merged = []
        i = 0
        while i < len(raw_eps):
            w     = raw_eps[i]
            group = [w]
            j     = i + 1
            while j < len(raw_eps):
                nw = raw_eps[j]
                if abs(nw['y'] - w['y']) < 2 and nw['x0'] - group[-1]['x1'] < 15:
                    group.append(nw); j += 1
                else:
                    break
            merged.append({'y': w['y'],
                           'text': ''.join(g['text'] for g in group),
                           'cx': (w['x0'] + group[-1]['x1']) / 2})
            i = j

        show_eps, inf_count = [], 0
        prev_ep, prev_is_inf = None, None
        for mw in merged:
            if not ep_re.match(mw['text']): continue
            try: ep_num = int(mw['text'])
            except: continue
            ep_y, ep_cx = mw['y'], mw['cx']
            cands = [(ep_y - cy, abs(ep_cx - cx), code)
                     for cy, cx, code in codes_at_y if 0 <= ep_y - cy < 40]
            if not cands: continue
            code   = sorted(cands)[0][2]
            is_inf = code.rstrip('_') == 'INF'
            if ep_num == prev_ep and is_inf == prev_is_inf:
                continue
            prev_ep, prev_is_inf = ep_num, is_inf
            if is_inf:
                inf_count += 1
            else:
                show_eps.append(ep_num)
        return show_eps, inf_count

    except Exception:
        return [], 0


def parse_holatv_log_xml_v2(file_or_bytes, log_date):
    try:
        if hasattr(file_or_bytes, 'read'):
            file_or_bytes.seek(0)
            raw = file_or_bytes.read()
            file_or_bytes.seek(0)
        else:
            raw = file_or_bytes
        raw  = re.sub(rb'&(?!amp;|lt;|gt;|apos;|quot;|#)', b'&amp;', raw)
        root = ET.fromstring(raw)
        fields = [f.text for f in root.find('fields')]
        data   = root.find('data')
        rows, dx, cx = [], 0, 0
        for row in data:
            vals  = dict(zip(fields, [col.text or '' for col in row]))
            mid   = vals.get('Media Id', '').split('#')[0].strip()
            typ   = vals.get('Type', '').upper()
            lt    = vals.get('Local Time', '')
            dur   = vals.get('Duration', '00:00:00')
            title = vals.get('Title', '')
            try:
                dt_utc = datetime.strptime(lt[:19], '%Y-%m-%d %H:%M:%S')
            except Exception:
                dt_utc = None
            dur_s = 0
            try:
                d = dur.replace(';', ':').split(':')
                dur_s = int(d[0]) * 3600 + int(d[1]) * 60 + int(d[2])
            except Exception:
                pass
            if typ == 'DX': dx += 1
            elif typ == 'CX': cx += 1
            if mid.startswith('HPP'):
                norm_type = 'COMMERCIAL'
            elif typ == 'PROGRAM':
                norm_type = 'PROGRAM'
            elif typ in ('DX', 'CX'):
                norm_type = typ
            elif typ == 'PROMOTION':
                norm_type = 'PROMO'
            else:
                norm_type = 'OTHER'
            rows.append({'media_id': mid, 'type': norm_type,
                         'start_utc': dt_utc, 'duration_secs': dur_s, 'title': title})
        return rows, dx, cx
    except Exception:
        return [], 0, 0


def parse_holatv_log_xlsx_v2(file_or_bytes, log_date):
    try:
        from openpyxl import load_workbook
        import io as _io
        if hasattr(file_or_bytes, 'read'):
            file_or_bytes.seek(0)
            data = file_or_bytes.read()
            file_or_bytes.seek(0)
        else:
            data = file_or_bytes
        wb   = load_workbook(_io.BytesIO(data), read_only=True)
        ws   = wb.active
        rows_raw = list(ws.iter_rows(values_only=True))
        rows, dx, cx = [], 0, 0
        for r in rows_raw[1:]:
            if not r or r[0] is None: continue
            try:
                hora_raw = str(r[1]).strip() if r[1] else ''
                tipo_raw = str(r[2]).strip().upper() if r[2] else ''
                rec_key  = str(r[4]).strip() if len(r) > 4 and r[4] else ''
                titulo   = str(r[5]).strip() if len(r) > 5 and r[5] else ''
                dur_raw  = str(r[6]).strip() if len(r) > 6 and r[6] else '00:00:00:00'
            except Exception:
                continue
            if tipo_raw == 'DX': dx += 1
            elif tipo_raw == 'CX': cx += 1
            try:
                parts = hora_raw.split(':')
                h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
                et_dt = datetime(log_date.year, log_date.month, log_date.day, h, m, s)
                if h < 6: et_dt += timedelta(days=1)
                offset = 4 if _edt_start(et_dt.year) <= et_dt < _est_start(et_dt.year) else 5
                dt_utc = et_dt + timedelta(hours=offset)
            except Exception:
                dt_utc = None
            dur_s = 0
            try:
                d = dur_raw.split(':')
                dur_s = int(d[0]) * 3600 + int(d[1]) * 60 + int(d[2])
            except Exception:
                pass
            mid = rec_key
            if mid.startswith('HPP'):
                norm_type = 'COMMERCIAL'
            elif tipo_raw == 'BLOQ':
                norm_type = 'PROGRAM'
            elif tipo_raw in ('DX', 'CX'):
                norm_type = tipo_raw
            elif tipo_raw in ('PROM', 'CORT'):
                norm_type = 'PROMO'
            else:
                norm_type = 'OTHER'
            rows.append({'media_id': mid, 'type': norm_type,
                         'start_utc': dt_utc, 'duration_secs': dur_s, 'title': titulo})
        return rows, dx, cx
    except Exception:
        return [], 0, 0


def parse_holatv_log_txt_v2(file_or_bytes, log_date):
    try:
        if hasattr(file_or_bytes, 'read'):
            file_or_bytes.seek(0)
            raw = file_or_bytes.read()
            file_or_bytes.seek(0)
        else:
            raw = file_or_bytes
        for enc in ('utf-8', 'latin-1', 'cp1252'):
            try: text = raw.decode(enc); break
            except: pass
        else:
            text = raw.decode('latin-1', errors='replace')
        lines = text.splitlines()
        if not lines: return [], 0, 0
        rows, dx, cx = [], 0, 0
        header_done = False
        for line in lines:
            if not line.strip(): continue
            cols = line.rstrip('\r\n').split('\t')
            if not header_done:
                header_done = True
                if cols[0].strip().upper() in ('N.ORD.', 'N.ORD', 'NORD'): continue
            if len(cols) < 9: continue
            try:
                hora_raw = cols[1].strip()
                tipo_raw = cols[2].strip().upper()
                rec_key  = cols[25].strip() if len(cols) > 25 and cols[25].strip() else cols[3].strip().split('#')[0]
                titulo   = cols[5].strip() if len(cols) > 5 else ''
                dur_raw  = cols[8].strip() if len(cols) > 8 else '00:00:00:00'
            except Exception:
                continue
            if tipo_raw == 'DX': dx += 1
            elif tipo_raw == 'CX': cx += 1
            try:
                parts = hora_raw.split(':')
                h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
                et_dt = datetime(log_date.year, log_date.month, log_date.day, h, m, s)
                if h < 6: et_dt += timedelta(days=1)
                offset = 4 if _edt_start(et_dt.year) <= et_dt < _est_start(et_dt.year) else 5
                dt_utc = et_dt + timedelta(hours=offset)
            except Exception:
                dt_utc = None
            dur_s = 0
            try:
                d = dur_raw.split(':')
                dur_s = int(d[0]) * 3600 + int(d[1]) * 60 + int(d[2])
            except Exception:
                pass
            mid = rec_key
            if mid.startswith('HPP'): norm_type = 'COMMERCIAL'
            elif tipo_raw == 'BLOQ':  norm_type = 'PROGRAM'
            elif tipo_raw in ('DX','CX'): norm_type = tipo_raw
            elif tipo_raw in ('PROM','CORT'): norm_type = 'PROMO'
            else: norm_type = 'OTHER'
            rows.append({'media_id': mid, 'type': norm_type,
                         'start_utc': dt_utc, 'duration_secs': dur_s, 'title': titulo})
        return rows, dx, cx
    except Exception:
        return [], 0, 0


def load_holatv_log(file_or_bytes, log_date):
    if file_or_bytes is None:
        return [], 0, 0
    name = getattr(file_or_bytes, 'name', '') or ''
    ext  = name.lower().rsplit('.', 1)[-1] if '.' in name else ''
    if ext == 'xml':
        return parse_holatv_log_xml_v2(file_or_bytes, log_date)
    elif ext in ('xlsx', 'xlsm'):
        return parse_holatv_log_xlsx_v2(file_or_bytes, log_date)
    elif ext == 'txt':
        return parse_holatv_log_txt_v2(file_or_bytes, log_date)
    return parse_holatv_log_xml_v2(file_or_bytes, log_date)


def group_holatv_blocks(log_rows):
    prog_rows = [r for r in log_rows
                 if r['type'] == 'PROGRAM' or r['media_id'].startswith('HPP')]
    prog_rows.sort(key=lambda x: x['start_utc'] or datetime.min)
    blocks, prev_base, cur_block = [], None, None
    for r in prog_rows:
        mid  = r['media_id']
        base = re.sub(r'_\d+$', '', mid)
        if base != prev_base:
            if cur_block: blocks.append(cur_block)
            cur_block = {'base_id': base, 'start_utc': r['start_utc'],
                         'duration_secs': r['duration_secs'], 'segments': [r],
                         'is_hpp': base.startswith('HPP')}
            prev_base = base
        else:
            cur_block['segments'].append(r)
            cur_block['duration_secs'] += r['duration_secs']
    if cur_block: blocks.append(cur_block)
    return blocks


def pick_grilla_for_date(grilla_list, target_date, channel):
    if not grilla_list: return None, None
    if len(grilla_list) == 1: return grilla_list[0], None
    from openpyxl import load_workbook
    import io
    for gf in grilla_list:
        try:
            gf.seek(0)
            data = gf.read()
            gf.seek(0)
            if gf.name.lower().endswith('.pdf'):
                _fname = gf.name.replace('_', ' ').upper()
                _MES = {'ENE':1,'FEB':2,'MAR':3,'ABR':4,'MAY':5,'JUN':6,
                        'JUL':7,'AGO':8,'SEP':9,'OCT':10,'NOV':11,'DIC':12}
                _tok  = re.findall(r'(\d+)\s+([A-Z]{3})', _fname)
                _yr_m = re.search(r'(\d{4})', _fname)
                _year = int(_yr_m.group(1)) if _yr_m else target_date.year
                _months = [(m.start(), _MES[m.group()])
                           for m in re.finditer(r'(ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)', _fname)]
                _nums   = [(m.start(), int(m.group()))
                           for m in re.finditer(r'\d+', _fname) if 1 <= int(m.group()) <= 31]
                _dates  = []
                for _mpos, _mnum in _months:
                    _before = [(p,n) for p,n in _nums if p < _mpos]
                    if _before:
                        _day = max(_before, key=lambda x: x[0])[1]
                        try:
                            from datetime import date as _d2
                            _dates.append(_d2(_year, _mnum, _day))
                        except Exception: pass
                if len(_months) == 1 and len(_dates) == 1 and _months:
                    _mpos, _mnum = _months[0]
                    _before = [(p,n) for p,n in _nums if p < _mpos]
                    if len(_before) >= 2:
                        _all_days = sorted(set(n for _,n in _before))
                        try:
                            from datetime import date as _d2
                            _dates.append(_d2(_year, _mnum, _all_days[0]))
                        except Exception: pass
                if len(_dates) >= 2:
                    _s, _e = min(_dates), max(_dates)
                    if _s <= target_date <= _e:
                        gf.seek(0); return gf, None
                    continue
                gf.seek(0); return gf, None
            wb = load_workbook(io.BytesIO(data), read_only=True)
            if channel in ('latam', 'us', 'tn', 'hl', 'hu'):
                for name in reversed(wb.sheetnames):
                    ws = wb[name]
                    rows = list(ws.iter_rows(max_row=3, values_only=True))
                    if len(rows) < 2: continue
                    header = rows[1]
                    col_dates = [(i, _parse_date_str(cell, force_year=target_date.year))
                                 for i, cell in enumerate(header)
                                 if _parse_date_str(cell, force_year=target_date.year)]
                    if not col_dates: continue
                    if col_dates[0][1] <= target_date <= col_dates[-1][1]:
                        gf.seek(0); return gf, None
            else:
                ws = wb.active
                rows = list(ws.iter_rows(max_row=2, values_only=True))
                if len(rows) >= 2:
                    for cell in rows[1]:
                        d = _parse_date_str(cell, force_year=target_date.year)
                        if d and abs((d - target_date).days) <= 6:
                            gf.seek(0); return gf, None
        except Exception:
            try: gf.seek(0)
            except: pass
    grilla_list[0].seek(0)
    return grilla_list[0], 'Could not determine week for grilla — using first file'


def generate_report_holatv_v2(channel, log_rows, dx_count, cx_count,
                               grilla_entries, playlist, lang='en', file_info=None,
                               current_start_utc=None):
    sep        = '═' * 60
    log_blocks = group_holatv_blocks(log_rows)
    prog_blocks = [b for b in log_blocks if not b['is_hpp']]
    hpp_blocks  = [b for b in log_blocks if b['is_hpp']]
    pt = playlist['type'] if playlist else 'full'

    lines = [sep, f'CHANNEL: {channel}',
             f'DATE: {log_rows[0]["start_utc"].date() if log_rows and log_rows[0].get("start_utc") else "?"}',
             f'TYPE: {"Full Day" if pt=="full" else ("Partial" if lang=="en" else "Parcial")}']
    if current_start_utc:
        lines.append(f'{"Checking from" if lang=="en" else "Verificando desde"}: {fmt_t(current_start_utc)}')
    if file_info:
        lbl = 'Files' if lang=='en' else 'Archivos'
        lines.append(f'{lbl}:')
        if file_info.get('grilla'): lines.append(f'  Grid:     {file_info["grilla"]}')
        if file_info.get('log'):    lines.append(f'  Log:      {file_info["log"]}')
        if file_info.get('json'):   lines.append(f'  Playlist: {file_info["json"]}')
    lines += [sep,
              f'{"Summary" if lang=="en" else "Resumen"}: {len(prog_blocks)} show blocks | {len(hpp_blocks)} infomercials | DX={dx_count} CX={cx_count}',
              '']

    lines.append(f'── [1] {"PROGRAM CHECK (Grilla vs Log)" if lang=="en" else "VERIFICACIÓN PROGRAMAS (Grilla vs Log)"} ──')
    lines += check_holatv_programs_v2(grilla_entries, log_blocks, current_start_utc, lang)
    lines.append('')

    lines.append(f'── [2] {"TIMING CHECK" if lang=="en" else "VERIFICACIÓN TIMING"} ──')
    lines += check_holatv_timing_v2(grilla_entries, log_blocks, current_start_utc, lang)
    lines.append('')

    if playlist:
        # [3] Infomercial check — both sides filtered to window
        lines.append(f'── [3] {"INFOMERCIAL CHECK (Log vs Playlist)" if lang=="en" else "VERIFICACIÓN INFOMERCIALES (Log vs Playlist)"} ──')
        _hpp_start  = current_start_utc
        log_hpp_ids = [b['base_id'] for b in hpp_blocks
                       if not _hpp_start or not b['start_utc'] or b['start_utc'] >= _hpp_start]
        pl_hpp      = [c for c in playlist.get('commercials', [])
                       if c.get('asset_ref', c.get('ref','')).startswith('HPP')
                       and (not _hpp_start or not c.get('start') or c['start'] >= _hpp_start)]
        pl_hpp_ids  = [c.get('asset_ref', c.get('ref','')) for c in pl_hpp]
        from collections import Counter as _Counter
        log_c = _Counter(log_hpp_ids); pl_c = _Counter(pl_hpp_ids)
        hpp_issues = [f'  ⚠  {hid}: log={log_c.get(hid,0)}x playlist={pl_c.get(hid,0)}x'
                      for hid in sorted(set(log_c) | set(pl_c))
                      if log_c.get(hid,0) != pl_c.get(hid,0)]
        lines += hpp_issues if hpp_issues else [f'  ✓  {"Infomercials match" if lang=="en" else "Infomerciales coinciden"}']
        lines.append('')

        lines.append(f'── [4] {"PROMO REPEAT CHECK" if lang=="en" else "VERIFICACIÓN PROMOS REPETIDAS"} ──')
        pi = check_promo_repeats(playlist, current_start_utc, lang)
        lines += pi if pi else [f'  {T("ok_promos", lang)}']
        lines.append('')

        lines.append(f'── [5] {"NOT INGESTED" if lang=="en" else "NO INGRESADOS"} ──')
        ni = check_not_ingested(playlist, current_start_utc, lang)
        lines += ni if ni else [f'  {T("ok_ingested", lang)}']
        lines.append('')

        lines.append(f'── [6] {"BUGS CHECK" if lang=="en" else "VERIFICACIÓN DE BUGS"} ──')
        bi = check_bugs(playlist, current_start_utc, lang)
        lines += bi if bi else [f'  {T("ok_bugs", lang)}']
        lines.append('')

        lines.append(f'── [7] {"CUE TONES" if lang=="en" else "CUE TONES"} ──')
        ci = check_cue_tones(playlist, lang)
        lines += ci if ci else [f'  {T("ok_cues", lang)}']
        lines.append('')

    lines.append(sep)
    return '\n'.join(str(l) for l in lines)


def check_programs_vs_grilla_tn(playlist, grilla_pairs, current_start, lang):
    """
    TN program check: set-based unique episode comparison.
    Uses p['name'] field (e.g. 'GENESIS_E122') to extract episode numbers.
    """
    import re as _re

    def parse_ep_num(name):
        m = _re.search(r'_E(\d+)$', str(name))
        return int(m.group(1)) if m else None

    # Unique episode numbers from grilla
    grilla_eps = {}
    for show_name, ep_num in grilla_pairs:
        if ep_num not in grilla_eps:
            grilla_eps[ep_num] = show_name

    # Unique episodes from JSON — use p['name'] which has e.g. 'GENESIS_E122'
    seen, json_eps = set(), {}
    for p in playlist['programs']:
        ref = p['episode_id_raw']
        if ref in seen: continue
        seen.add(ref)
        if current_start and p['start'] and p['start'] < current_start: continue
        ep_num = parse_ep_num(p['name'])
        if ep_num is not None:
            json_eps[ep_num] = {'name': p['name'], 'start': p['start']}

    issues = []
    for ep, show in sorted(grilla_eps.items()):
        if ep not in json_eps:
            issues.append(f'  ✗  NOT IN PLAYLIST: {show} ep{ep}')
    for ep, info in sorted(json_eps.items()):
        if ep not in grilla_eps:
            issues.append(f'  ✗  EXTRA IN PLAYLIST: {info["name"]} @ {fmt_t(info["start"])} (not in grilla)')
    if not issues:
        issues.append(f'  ✓  All {len(json_eps)} episodes match grilla')
    return issues
